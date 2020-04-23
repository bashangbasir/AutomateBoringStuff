# workbook - excel file
# workbook can have multiple sheet

import openpyxl
import os 

# path of working dir
wd = "F:\\AutomateBoringStuff\\WordExcelPDFDocument\\Excel"
os.chdir(wd)

# load workbook using load_workbook() method, return workbook object
workbook = openpyxl.load_workbook("example.xlsx")
print(workbook)

# to get sheet in workbook, use get_sheet_by_name() method in workbook object, 
# this will return sheet object
# however this method is deprecate during this time
# can use workbook["Sheet1"] to get the same thing. 

sheet = workbook["Sheet1"]
sheet2 = workbook.get_sheet_by_name("Sheet2")#this will show error! since this method is deprecated.
print(sheet)
print(sheet2) 

# to get sheet name, use get_sheet_name() method

sheet_name = workbook.get_sheet_names()

print(sheet_name)

# to get cell from sheet,
cell = sheet["A1"] #this will return cell object 
#to get the value of the specific cell 
cell_value = cell.value
print(cell_value)
#the type cell value refer to the calue in the cell, for this cell it is in datetime object 
#since the cell value is a date 
print(type(cell_value))

#also can use this format to get cell 
cell2 = sheet.cell(row=1,column=2) # this same as : sheet["B1"]
cell2_value = cell2.value
print(cell2_value)

#this method we can make easier way to scrap the cell,
#example
for i in range(1,8):
    print(i, sheet.cell(row=i,column=2).value)

