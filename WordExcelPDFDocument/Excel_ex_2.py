import openpyxl
import os

cwd = os.chdir(r"F:\AutomateBoringStuff\WordExcelPDFDocument\Excel")

# create Workbook - this will return workbook object
wb = openpyxl.Workbook()# this code give same output as line 16

#sheet_name = wb.get_sheet_names()
sheet_name = wb.sheetnames
print(sheet_name)

# sheet = wb.get_sheet_by_name("Sheet")
sheet = wb["Sheet"] # this code give same output as line 12

# assign value to cell is just like assign value to variable
if sheet["A1"].value == None: # if statement to check if the cell is already occupied or empty
    sheet["A1"] = 42 # cell A1 is assigned with 42

if sheet["A2"].value == None: # if statement to check if the cell is already occupied or empty
    sheet["A2"] = "Hello"

# to create new sheet  
# this will add sheet at the end of sheet list
sheet2 = wb.create_sheet()

# when index is sent to the method, it will position the sheet according to the index value. start at 0.
sheet3 = wb.create_sheet(index=2)

#check sheet will be 3 since we create new one
sheet_name = wb.sheetnames
print(sheet_name)

#to set title for sheet 
sheet.title = "My first sheet"
sheet2.title = "Second sheet"
sheet3.title = "Third sheet"

sheet_name = wb.sheetnames
print(sheet_name)

# to save the workbook 
wb.save("CreateExcel.xlsx")

# to load back the workbook 
path = "Excel path"
wb1 = openpyxl.load_workbook(path)